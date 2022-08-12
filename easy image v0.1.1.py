import os
import openpyxl
from reportlab.lib import colors
from reportlab.graphics.shapes import *
from reportlab.graphics import renderPDF

def printfirst(sheet):
    sum=[]
    print('|',end=' ')
    for cell in sheet['1']:
        sum.append(cell.value)
        print(cell.value,end =' | ')
    else: print()
def readfirst(sheet):
    sum= []
    for cell in sheet['1']:
        sum.append(cell.value)
    return sum
#重写为类

path =r'C:\Users\deepwind\Desktop'    # 配置1，工作目录
os.chdir(path)

workbook = openpyxl.load_workbook(input('输入文件名：'))   #配置2 对象表格 
print('全部工作表：', workbook.sheetnames)
for i in workbook.sheetnames:
    print(i,'表大小：', workbook[i].dimensions,end='    ') # 表尺寸展示
    print('表预览:',end='')
    printfirst(workbook[i]) # 表预览

print('活动表：', workbook.active)
choicesheet= input('选择你要使用的工作表：')
#KeyError
sheet = workbook[choicesheet]
#要打印表数据吗？
print('请选择您需要的数据：')
print('|',end=' ')
printfirst(sheet)

row=readfirst(sheet)
data = {}.fromkeys(row)

#data['time']

#获取输入
input0=input().split()
#while True:
input0.append('当地时间')

#导入数据到data
for i in input0:
    if i in row:
        sum0=[]
        for j in sheet[chr(row.index(i)+65)]:
            sum0.append(j.value)
        else:
            sum0.pop(0)
            data[i]=sum0

#时间处理到小时
month = [31,28,31,30,31,30,31,31,30,31,30,31]
sum,l=0,[]
for i in month:
    l.append(sum)
    sum+=i
hours=[]
for i in data['当地时间']:
    ti=i[:-3].replace('.',' ')
    hour=int(ti[:2])*24+l[int(ti[3:5])-1]*24+int(ti[-3:])
    hours.append(hour)
del data['当地时间'] #节约内存
if hours[0]>hours[1]:
    hours.reverse()
    for i in data.items():
        if i[1]:
            i[1].reverse()

#开始绘制
drawing=Drawing(2000,200)
for i in input0:
    if i != '当地时间':
        drawing.add(PolyLine(list(zip(hours,data[i]))))
        #关于None的错误处理
        #关于风向等的错误处理
        #输出随数据而改变的合理坐标图

renderPDF.drawToFile(drawing,'report1.pdf','Weather')













